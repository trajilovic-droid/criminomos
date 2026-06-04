# Criminomos - Serveur MCP v13
# Compatible Claude.ai avec endpoints OAuth 2.1 discovery
import json
import re
import unicodedata
import os
import uuid
import logging
import io
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timezone

import httpx
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from starlette.applications import Starlette
from starlette.requests import Request
from starlette.responses import JSONResponse, Response
from starlette.routing import Route
from starlette.middleware import Middleware
from starlette.middleware.cors import CORSMiddleware
import uvicorn

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
GDRIVE_FILE_ID  = os.environ.get("GDRIVE_FILE_ID", "18ylKTce78zSdEpeJ8tBbPchPIGs-kG4w")
RELOAD_KEY      = os.environ.get("RELOAD_KEY", "iuris2026!")
GDRIVE_URL      = f"https://docs.google.com/spreadsheets/d/{GDRIVE_FILE_ID}/export?format=xlsx"
BASE_URL        = os.environ.get("BASE_URL", "https://mcp.criminomos.ch")
MAX_SESSIONS    = int(os.environ.get("MAX_SESSIONS_PER_TOKEN", "2"))
MAX_LOG_ENTRIES = 500

def load_tokens():
    raw = os.environ.get("ACCESS_TOKENS", "")
    if not raw.strip():
        return set()
    return {t.strip() for t in raw.split(",") if t.strip()}

ACCESS_TOKENS = load_tokens()

# ---------------------------------------------------------------------------
# Journal d'accès
# ---------------------------------------------------------------------------
ACCESS_LOG = []

def log_access(token, ip, action):
    entry = {
        "token":     token,
        "ip":        ip,
        "action":    action,
        "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    }
    ACCESS_LOG.append(entry)
    if len(ACCESS_LOG) > MAX_LOG_ENTRIES:
        ACCESS_LOG.pop(0)
    logger.info(f"ACCESS | token={token} | ip={ip} | action={action}")

TOKEN_SESSIONS = defaultdict(set)

# ---------------------------------------------------------------------------
# Authentification
# ---------------------------------------------------------------------------
def get_token(request: Request) -> str:
    auth = request.headers.get("authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:].strip()
    return request.headers.get("x-api-key", "").strip()

def is_authorized(request: Request) -> tuple:
    if not ACCESS_TOKENS:
        return True, "anonymous", "ok"
    token = get_token(request)
    if not token:
        return False, "", "Token manquant"
    if token not in ACCESS_TOKENS:
        return False, token, "Token invalide"
    return True, token, "ok"

def check_concurrent_sessions(token: str, session_id: str) -> bool:
    sessions = TOKEN_SESSIONS[token]
    if session_id in sessions:
        return True
    if len(sessions) >= MAX_SESSIONS:
        return False
    return True

# ---------------------------------------------------------------------------
# Dictionnaire multilingue
# ---------------------------------------------------------------------------
MULTILANG = {
    "expulsion": ["landesverweisung", "espulsione"],
    "viol": ["vergewaltigung", "violenza carnale"],
    "meurtre": ["mord", "omicidio"],
    "homicide": ["tötung", "omicidio"],
    "lésions corporelles": ["körperverletzung", "lesioni corporali"],
    "escroquerie": ["betrug", "truffa"],
    "abus de confiance": ["veruntreuung", "appropriazione indebita"],
    "contrainte": ["nötigung", "coazione"],
    "menaces": ["drohung", "minaccia"],
    "brigandage": ["raub", "rapina"],
    "vol": ["diebstahl", "furto"],
    "faux": ["urkundenfälschung", "falsità in documenti"],
    "diffamation": ["verleumdung", "diffamazione"],
    "injure": ["beschimpfung", "ingiuria"],
    "incendie": ["brandstiftung", "incendio"],
    "tentative": ["versuch", "tentativo"],
    "complicité": ["gehilfenschaft", "complicità"],
    "instigation": ["anstiftung", "istigazione"],
    "récidive": ["rückfall", "recidiva"],
    "concours": ["konkurrenz", "concorso"],
    "détention provisoire": ["untersuchungshaft", "carcerazione preventiva"],
    "détention": ["haft", "detenzione"],
    "arrestation": ["verhaftung", "arresto"],
    "ordonnance pénale": ["strafbefehl", "decreto d'accusa"],
    "classement": ["einstellung", "abbandono"],
    "non-entrée en matière": ["nichtanhandnahme", "non luogo a procedere"],
    "acquittement": ["freispruch", "assoluzione"],
    "condamnation": ["verurteilung", "condanna"],
    "appel": ["berufung", "appello"],
    "recours": ["beschwerde", "ricorso"],
    "révision": ["revision", "revisione"],
    "récusation": ["ausstand", "ricusazione"],
    "scellés": ["siegelung", "sigillazione"],
    "séquestre": ["beschlagnahme", "sequestro"],
    "perquisition": ["hausdurchsuchung", "perquisizione"],
    "surveillance": ["überwachung", "sorveglianza"],
    "expertise": ["gutachten", "perizia"],
    "peine privative de liberté": ["freiheitsstrafe", "pena detentiva"],
    "peine pécuniaire": ["geldstrafe", "pena pecuniaria"],
    "sursis": ["aufschub", "sospensione"],
    "libération conditionnelle": ["bedingte entlassung", "liberazione condizionale"],
    "internement": ["verwahrung", "internamento"],
    "mesure": ["massnahme", "misura"],
    "culpabilité": ["schuld", "colpevolezza"],
    "négligence": ["fahrlässigkeit", "negligenza"],
    "causalité": ["kausalität", "causalità"],
    "prescription": ["verjährung", "prescrizione"],
    "légitime défense": ["notwehr", "legittima difesa"],
    "état de nécessité": ["notstand", "stato di necessità"],
    "victime": ["opfer", "vittima"],
    "prévenu": ["beschuldigte", "imputato"],
    "ministère public": ["staatsanwaltschaft", "ministero pubblico"],
}

def expand_query_multilang(query_words):
    expanded    = list(query_words)
    query_lower = " ".join(query_words).lower()
    for fr_term, translations in MULTILANG.items():
        if fr_term in query_lower:
            expanded.extend(translations)
    return expanded

# ---------------------------------------------------------------------------
# Chargement des données
# ---------------------------------------------------------------------------
def parse_excel(content: bytes) -> list:
    sheets  = pd.read_excel(io.BytesIO(content), sheet_name=None)
    wb      = load_workbook(io.BytesIO(content))
    url_map = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink and cell.value:
                    url_map[str(cell.value).strip()] = cell.hyperlink.target
    all_rows = []
    for sheet_name, df in sheets.items():
        if sheet_name == "2021-2024":
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
        for _, row in df.iterrows():
            r = {}
            for col in df.columns:
                val = row.get(col, None)
                if pd.notna(val):
                    r[str(col).strip()] = str(val)
            all_rows.append(r)
    trimmed = []
    for r in all_rows:
        arret    = r.get("Arrêt", "").strip()
        parution = r.get("Date de parution", "")
        decision = r.get("Date de la décision", "")
        trimmed.append({
            "arret":    arret,
            "parution": parution.split(" ")[0] if "00:00:00" in parution else parution,
            "decision": decision.split(" ")[0] if "00:00:00" in decision else decision,
            "objet":    r.get("Objet", ""),
            "articles": r.get("Articles", ""),
            "resume":   r.get("Résumé", ""),
            "langue":   r.get("Langue", "").strip() if r.get("Langue") else "",
            "interet":  r.get("Arrêt d'intérêt", "").strip() if r.get("Arrêt d'intérêt") else r.get("Arrêt d intérêt", ""),
            "admis":    r.get("Admis/rejeté", ""),
            "peine":    r.get("Peine prononcée", ""),
            "url":      url_map.get(arret, ""),
        })
    return [r for r in trimmed if r["arret"]]


def load_from_gdrive():
    logger.info("Téléchargement depuis Google Drive...")
    resp = httpx.get(GDRIVE_URL, timeout=60, follow_redirects=True)
    resp.raise_for_status()
    data  = parse_excel(resp.content)
    by_id = {r["arret"]: r for r in data}
    logger.info(f"=== {len(data)} arrêts chargés ===")
    return data, by_id


def load_from_local():
    local = Path(__file__).parent / "arrets.json"
    if local.exists():
        with open(local, encoding="utf-8") as f:
            data = json.load(f)
        by_id = {r["arret"]: r for r in data if r.get("arret")}
        logger.info(f"=== {len(data)} arrêts chargés (local) ===")
        return data, by_id
    return [], {}


try:
    ARRETS, ARRETS_BY_ID = load_from_gdrive()
except Exception as e:
    logger.warning(f"Google Drive inaccessible ({e}), chargement local.")
    ARRETS, ARRETS_BY_ID = load_from_local()

SESSIONS = {}

# ---------------------------------------------------------------------------
# Outils MCP
# ---------------------------------------------------------------------------
TOOLS = [
    {
        "name": "search_arrets",
        "description": "Recherche des arrets du Tribunal federal suisse en droit penal. Recherche automatiquement dans les trois langues (FR/DE/IT).",
        "inputSchema": {
            "type": "object",
            "properties": {
                "query":      {"type": "string",  "description": "Mots-cles de recherche"},
                "infraction": {"type": "string",  "description": "Type d infraction ex: Expulsion"},
                "article":    {"type": "string",  "description": "Article de loi ex: 66a CP"},
                "annee":      {"type": "string",  "description": "Annee ex: 2024"},
                "langue":     {"type": "string",  "description": "Langue : F, D ou I"},
                "limite":     {"type": "integer", "description": "Nombre de resultats max 30"}
            },
            "required": ["query"]
        }
    },
    {
        "name": "get_fulltext",
        "description": "Charge le texte integral d un arret depuis bger.ch.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "arret_id": {"type": "string", "description": "Numero d arret ex: 6B_409/2024"}
            },
            "required": ["arret_id"]
        }
    },
    {
        "name": "get_references",
        "description": "Extrait les ATF et arrets TF cites dans un arret (niveau 1).",
        "inputSchema": {
            "type": "object",
            "properties": {
                "arret_id": {"type": "string", "description": "Numero d arret ex: 6B_409/2024"}
            },
            "required": ["arret_id"]
        }
    },
    {
        "name": "get_references_deep",
        "description": "Remonte les references sur 2 niveaux de profondeur.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "arret_id": {"type": "string",  "description": "Numero d arret de depart"},
                "max_refs": {"type": "integer", "description": "Nombre max de refs niveau 1 (defaut 5, max 10)"}
            },
            "required": ["arret_id"]
        }
    },
    {
        "name": "get_arret_by_reference",
        "description": "Charge le texte d un ATF ou arret TF cite en reference.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "reference": {"type": "string", "description": "ex: ATF 148 IV 409 ou 6B_123/2021"}
            },
            "required": ["reference"]
        }
    }
]

# ---------------------------------------------------------------------------
# Logique métier
# ---------------------------------------------------------------------------
def normalize(text):
    text = text.lower()
    text = unicodedata.normalize("NFD", text)
    return "".join(c for c in text if unicodedata.category(c) != "Mn")


def score_arret(arret, query_words):
    score = 0
    for word in query_words:
        w = normalize(word)
        if w in normalize(arret.get("objet",    "")): score += 4
        if w in normalize(arret.get("resume",   "")): score += 2
        if w in normalize(arret.get("articles", "")): score += 3
        if w in normalize(arret.get("arret",    "")): score += 5
    if arret.get("interet") == "oui":
        score += 1
    return score


def search_arrets(query="", infraction="", article="", annee="", langue="", limite=10):
    limite      = min(int(limite), 30)
    query_words = query.strip().split() if query.strip() else []
    expanded    = expand_query_multilang(query_words)
    results     = []
    for arret in ARRETS:
        if infraction and not normalize(arret.get("objet", "")).startswith(normalize(infraction)):
            continue
        if article and normalize(article) not in normalize(arret.get("articles", "")):
            continue
        if annee and not (arret.get("decision", "").startswith(annee) or
                          arret.get("parution",  "").startswith(annee)):
            continue
        if langue and arret.get("langue", "").upper() != langue.upper():
            continue
        score = score_arret(arret, expanded) if expanded else 1
        if score > 0 or not query_words:
            results.append((score, arret))
    results.sort(key=lambda x: x[0], reverse=True)
    results = results[:limite]
    if not results:
        return "Aucun arret trouve."
    extra  = [w for w in expanded if w not in query_words]
    header = ("Recherche etendue aux equivalents : " + ", ".join(extra) + "\n\n") if extra else ""
    lines  = [header + str(len(results)) + " arret(s) trouve(s)\n"]
    for _, r in results:
        lines.append("### " + r["arret"])
        lines.append("- Objet : "    + r.get("objet",    "-"))
        lines.append("- Date : "     + r.get("decision", "-"))
        lines.append("- Langue : "   + r.get("langue",   "-"))
        lines.append("- Articles : " + r.get("articles", "-"))
        if r.get("admis"):             lines.append("- Resultat : " + r["admis"])
        if r.get("interet") == "oui": lines.append("- Arret d interet")
        if r.get("resume"):            lines.append("- Resume : " + r["resume"])
        if r.get("url"):               lines.append("- URL : " + r["url"])
        lines.append("")
    return "\n".join(lines)


def _fetch_text(url):
    resp = httpx.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"}, follow_redirects=True)
    soup = BeautifulSoup(resp.text, "html.parser")
    for tag in soup(["script", "style", "nav", "header", "footer"]):
        tag.decompose()
    lines = [l.rstrip() for l in soup.get_text(separator="\n").splitlines() if l.strip()]
    return "\n".join(lines)


def _extract_refs(text, arret_id=""):
    atf_refs = sorted(set(re.findall(r"ATF\s+\d{2,3}\s+[IVX]+\s+\d+", text)))
    tf_refs  = sorted(set(
        r for r in re.findall(r"\b[0-9][A-Z]{1,2}_\d{1,4}/20\d{2}\b", text)
        if r != arret_id
    ))
    return atf_refs, tf_refs


def _arret_url(arret_id):
    arret = ARRETS_BY_ID.get(arret_id)
    if arret and arret.get("url"):
        return arret["url"]
    cid = re.sub(r"[^A-Za-z0-9_/.-]", "", arret_id)[:40]
    return ("https://www.bger.ch/ext/eurospider/live/fr/php/aza/http/index.php"
            f"?lang=fr&type=show_document&highlight_docid=aza://{cid}")


def get_fulltext(arret_id):
    url = _arret_url(arret_id)
    try:
        text = _fetch_text(url)
        return "Arret " + arret_id + "\nURL : " + url + "\n" + "-"*60 + "\n\n" + text[:15000]
    except Exception as e:
        return "Erreur : " + str(e)


def get_references(arret_id):
    url = _arret_url(arret_id)
    try:
        text = _fetch_text(url)
    except Exception as e:
        return "Erreur : " + str(e)
    atf_refs, tf_refs = _extract_refs(text, arret_id)
    lines = ["References de l arret " + arret_id + "\n"]
    if atf_refs:
        lines.append("ATF cites (" + str(len(atf_refs)) + ") :")
        for r in atf_refs: lines.append("- " + r)
    if tf_refs:
        lines.append("\nArrets TF cites (" + str(len(tf_refs)) + ") :")
        for r in tf_refs:
            if r in ARRETS_BY_ID:
                a = ARRETS_BY_ID[r]
                lines.append("- " + r + " - " + a.get("objet", "-") +
                              " (" + a.get("decision", "-") + ") [base]")
            else:
                lines.append("- " + r)
    if not atf_refs and not tf_refs:
        lines.append("Aucune reference trouvee.")
    return "\n".join(lines)


def get_references_deep(arret_id, max_refs=5):
    max_refs = min(int(max_refs), 10)
    url      = _arret_url(arret_id)
    lines    = ["=== REFERENCES PROFONDES : " + arret_id + " ===\n"]
    try:
        text1 = _fetch_text(url)
    except Exception as e:
        return "Erreur : " + str(e)
    atf1, tf1 = _extract_refs(text1, arret_id)
    lines.append("NIVEAU 1 — References directes")
    lines.append("ATF cites : "       + (", ".join(atf1)     if atf1 else "aucun"))
    lines.append("Arrets TF cites : " + (", ".join(tf1[:20]) if tf1  else "aucun"))
    lines.append("")
    for ref_id in tf1[:max_refs]:
        try:
            text2     = _fetch_text(_arret_url(ref_id))
            atf2, tf2 = _extract_refs(text2, ref_id)
            meta      = ARRETS_BY_ID.get(ref_id)
            objet     = meta.get("objet", "—") if meta else "hors base"
            lines.append("\n  " + ref_id + " (" + objet + ")")
            if atf2: lines.append("  ATF cites : "       + ", ".join(atf2[:5]))
            if tf2:  lines.append("  Arrets TF cites : " + ", ".join(tf2[:5]))
            if not atf2 and not tf2: lines.append("  (aucune reference)")
        except Exception as e:
            lines.append("\n  " + ref_id + " — Erreur : " + str(e))
    for atf_ref in atf1[:3]:
        m = re.match(r"ATF\s+(\d{2,3})\s+([IVX]+)\s+(\d+)", atf_ref)
        if m:
            vol, part, page = m.groups()
            atf_url = ("https://www.bger.ch/ext/eurospider/live/fr/php/clir/http/index.php"
                       f"?lang=fr&type=show_document&highlight_docid=atf:///{vol}/{part}/{page}")
            try:
                text_atf  = _fetch_text(atf_url)
                atf2, tf2 = _extract_refs(text_atf, "")
                lines.append("\n  " + atf_ref)
                if atf2: lines.append("  ATF cites : "       + ", ".join(atf2[:5]))
                if tf2:  lines.append("  Arrets TF cites : " + ", ".join(tf2[:5]))
            except Exception as e:
                lines.append("\n  " + atf_ref + " — Erreur : " + str(e))
    return "\n".join(lines)


def get_arret_by_reference(reference):
    reference = reference.strip()
    if re.match(r"^[0-9][A-Z]{1,2}_\d{1,4}/20\d{2}$", reference):
        return get_fulltext(reference)
    m = re.match(r"ATF\s+(\d{2,3})\s+([IVX]+)\s+(\d+)", reference, re.IGNORECASE)
    if m:
        vol, part, page = m.groups()
        url = ("https://www.bger.ch/ext/eurospider/live/fr/php/clir/http/index.php"
               f"?lang=fr&type=show_document&highlight_docid=atf:///{vol}/{part}/{page}")
        try:
            text = _fetch_text(url)
            return reference + "\nURL : " + url + "\n" + "-"*60 + "\n\n" + text[:15000]
        except Exception as e:
            return "Erreur : " + str(e)
    return "Format non reconnu : " + reference


def call_tool(name, args):
    if name == "search_arrets":            return search_arrets(**args)
    elif name == "get_fulltext":           return get_fulltext(**args)
    elif name == "get_references":         return get_references(**args)
    elif name == "get_references_deep":    return get_references_deep(**args)
    elif name == "get_arret_by_reference": return get_arret_by_reference(**args)
    return "Outil inconnu : " + name

# ---------------------------------------------------------------------------
# Handlers HTTP
# ---------------------------------------------------------------------------
def ok(req_id, result):
    return JSONResponse({"jsonrpc": "2.0", "id": req_id, "result": result},
                        headers={"Content-Type": "application/json"})

def err(req_id, code, msg):
    return JSONResponse({"jsonrpc": "2.0", "id": req_id, "error": {"code": code, "message": msg}},
                        headers={"Content-Type": "application/json"})


# ---------------------------------------------------------------------------
# OAuth 2.1 Discovery Endpoints (requis par Claude.ai)
# Ces endpoints indiquent à Claude.ai que le serveur est public (pas d'auth requise)
# ---------------------------------------------------------------------------
async def handle_oauth_protected_resource(request: Request):
    """RFC 9728 — Protected Resource Metadata.
    Indique à Claude.ai comment s'authentifier (ici : pas d'auth requise).
    """
    return JSONResponse({
        "resource":                  BASE_URL,
        "authorization_servers":     [BASE_URL],
        "bearer_methods_supported":  ["header"],
        "scopes_supported":          [],
    })


async def handle_oauth_authorization_server(request: Request):
    """RFC 8414 — Authorization Server Metadata.
    Serveur OAuth minimal qui accepte tous les tokens.
    """
    return JSONResponse({
        "issuer":                                BASE_URL,
        "authorization_endpoint":               BASE_URL + "/oauth/authorize",
        "token_endpoint":                        BASE_URL + "/oauth/token",
        "registration_endpoint":                 BASE_URL + "/oauth/register",
        "response_types_supported":              ["code"],
        "grant_types_supported":                 ["authorization_code"],
        "code_challenge_methods_supported":      ["S256"],
        "token_endpoint_auth_methods_supported": ["none"],
    })


async def handle_oauth_authorize(request: Request):
    """Endpoint d'autorisation OAuth — redirige directement vers Claude.ai."""
    redirect_uri  = request.query_params.get("redirect_uri", "https://claude.ai/api/mcp/auth_callback")
    state         = request.query_params.get("state", "")
    code          = str(uuid.uuid4()).replace("-", "")[:16]
    separator     = "&" if "?" in redirect_uri else "?"
    redirect_url  = f"{redirect_uri}{separator}code={code}&state={state}"
    return Response(
        status_code=302,
        headers={"Location": redirect_url}
    )


async def handle_oauth_token(request: Request):
    """Endpoint de token OAuth — émet un token anonyme."""
    return JSONResponse({
        "access_token":  "criminomos-public-" + str(uuid.uuid4())[:8],
        "token_type":    "Bearer",
        "expires_in":    86400,
        "scope":         "",
    })


async def handle_oauth_register(request: Request):
    """Dynamic Client Registration — accepte tout client."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    client_id = "client-" + str(uuid.uuid4())[:8]
    return JSONResponse({
        "client_id":                client_id,
        "client_secret":            "",
        "redirect_uris":            body.get("redirect_uris", []),
        "grant_types":              ["authorization_code"],
        "response_types":           ["code"],
        "token_endpoint_auth_method": "none",
    }, status_code=201)


# ---------------------------------------------------------------------------
# Handlers principaux
# ---------------------------------------------------------------------------
async def handle_health(request: Request):
    return JSONResponse({
        "status":   "ok",
        "name":     "criminomos",
        "arrets":   len(ARRETS),
        "version":  "13.0",
        "auth":     "enabled" if ACCESS_TOKENS else "disabled",
        "sessions": {t: len(s) for t, s in TOKEN_SESSIONS.items()}
    })


async def handle_reload(request: Request):
    global ARRETS, ARRETS_BY_ID
    key = request.query_params.get("key", "")
    if key != RELOAD_KEY:
        return JSONResponse({"error": "Clé invalide"}, status_code=403)
    try:
        ARRETS, ARRETS_BY_ID = load_from_gdrive()
        return JSONResponse({"status": "ok", "arrets": len(ARRETS)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


async def handle_log(request: Request):
    key = request.query_params.get("key", "")
    if key != RELOAD_KEY:
        return JSONResponse({"error": "Clé invalide"}, status_code=403)
    token_filter = request.query_params.get("token", "")
    logs = ACCESS_LOG if not token_filter else [e for e in ACCESS_LOG if e["token"] == token_filter]
    return JSONResponse({"total": len(logs), "entries": list(reversed(logs[-100:]))})


async def handle_revoke(request: Request):
    key   = request.query_params.get("key", "")
    token = request.query_params.get("token", "")
    if key != RELOAD_KEY:
        return JSONResponse({"error": "Clé invalide"}, status_code=403)
    if not token:
        return JSONResponse({"error": "Token manquant"}, status_code=400)
    count = len(TOKEN_SESSIONS.get(token, set()))
    TOKEN_SESSIONS[token] = set()
    log_access(token, request.client.host if request.client else "unknown", "revoked_by_admin")
    return JSONResponse({"status": "ok", "token": token, "sessions_revoked": count})


async def handle_mcp(request: Request):
    if request.method == "GET":
        return JSONResponse(
            {"name": "criminomos", "version": "13.0", "protocolVersion": "2025-11-25"},
            headers={"MCP-Protocol-Version": "2025-11-25"}
        )

    if request.method == "HEAD":
        return Response(
            status_code=200,
            headers={"MCP-Protocol-Version": "2025-11-25"}
        )

    if request.method == "DELETE":
        sid   = request.headers.get("mcp-session-id")
        token = SESSIONS.pop(sid, None) if sid else None
        if token and sid:
            TOKEN_SESSIONS[token].discard(sid)
            log_access(token, request.client.host if request.client else "unknown", "disconnect")
        return Response(status_code=200)

    # Vérification du token pour POST
    authorized, token, msg = is_authorized(request)
    if not authorized:
        log_access(token or "unknown", request.client.host if request.client else "unknown", "denied:" + msg)
        return JSONResponse({"error": msg}, status_code=401)

    try:
        data = await request.json()
    except Exception:
        return err(None, -32700, "Parse error")

    method = data.get("method", "")
    params = data.get("params", {})
    req_id = data.get("id", 1)
    ip     = request.client.host if request.client else "unknown"

    if method == "initialize":
        sid = str(uuid.uuid4())
        if not check_concurrent_sessions(token, sid):
            log_access(token, ip, "blocked:max_sessions")
            return JSONResponse(
                {"error": f"Limite de {MAX_SESSIONS} session(s) simultanée(s) atteinte."},
                status_code=429
            )
        SESSIONS[sid]             = token
        TOKEN_SESSIONS[token].add(sid)
        log_access(token, ip, "connect")
        resp = ok(req_id, {
            "protocolVersion": "2025-11-25",
            "capabilities":    {"tools": {"listChanged": False}},
            "serverInfo":      {"name": "criminomos", "version": "13.0"}
        })
        resp.headers["mcp-session-id"] = sid
        return resp

    if method == "notifications/initialized":
        return Response(status_code=202)

    if method == "ping":
        return ok(req_id, {})

    if method == "tools/list":
        log_access(token, ip, "tools/list")
        return ok(req_id, {"tools": TOOLS})

    if method == "tools/call":
        tool = params.get("name", "")
        args = params.get("arguments", {})
        log_access(token, ip, "call:" + tool)
        try:
            result = call_tool(tool, args)
        except Exception as e:
            result = "Erreur : " + str(e)
        return ok(req_id, {"content": [{"type": "text", "text": result}]})

    return err(req_id, -32601, "Method not found")


# ---------------------------------------------------------------------------
# Application
# ---------------------------------------------------------------------------
middleware = [
    Middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_methods=["GET", "POST", "DELETE", "HEAD", "OPTIONS"],
        allow_headers=["*"],
        expose_headers=["mcp-session-id", "MCP-Protocol-Version"]
    )
]

app = Starlette(
    routes=[
        # Santé & admin
        Route("/",        handle_health,  methods=["GET", "HEAD"]),
        Route("/reload",  handle_reload,  methods=["GET"]),
        Route("/log",     handle_log,     methods=["GET"]),
        Route("/revoke",  handle_revoke,  methods=["GET"]),
        # OAuth 2.1 discovery (requis par Claude.ai)
        Route("/.well-known/oauth-protected-resource",   handle_oauth_protected_resource,   methods=["GET"]),
        Route("/.well-known/oauth-authorization-server", handle_oauth_authorization_server, methods=["GET"]),
        Route("/oauth/authorize", handle_oauth_authorize, methods=["GET"]),
        Route("/oauth/token",     handle_oauth_token,     methods=["POST"]),
        Route("/oauth/register",  handle_oauth_register,  methods=["POST"]),
        # MCP
        Route("/mcp", handle_mcp, methods=["GET", "POST", "DELETE", "HEAD"]),
    ],
    middleware=middleware
)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="info")
